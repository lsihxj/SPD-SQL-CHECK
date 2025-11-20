"""
导出服务
支持Excel和PDF格式导出
"""
import io
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from sqlalchemy.orm import Session

from app.models.database import SQLCheckRecord, CheckSummary


class ExportService:
    """导出服务"""
    
    def __init__(self, db_session: Session):
        """
        初始化导出服务
        
        Args:
            db_session: 数据库会话
        """
        self.db = db_session
    
    def export_to_excel(self, batch_id: str) -> bytes:
        """
        导出为Excel文件
        
        Args:
            batch_id: 批次ID
            
        Returns:
            Excel文件字节流
        """
        # 获取批次汇总
        summary = self.db.query(CheckSummary).filter(
            CheckSummary.batch_id == batch_id
        ).first()
        
        if not summary:
            raise ValueError(f"Batch {batch_id} not found")
        
        # 获取所有检查记录
        records = self.db.query(SQLCheckRecord).filter(
            SQLCheckRecord.batch_id == batch_id
        ).order_by(SQLCheckRecord.id).all()
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "SQL检查报告"
        
        # 设置列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 15
        
        # 标题样式
        title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        title_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 写入标题
        ws.append(['编号', 'SQL语句', '检查状态', 'AI检查结果', '错误信息', '耗时(ms)'])
        for cell in ws[1]:
            cell.font = title_font
            cell.fill = title_fill
            cell.alignment = title_alignment
        
        # 写入数据
        for idx, record in enumerate(records, start=1):
            ws.append([
                idx,
                record.original_sql,
                record.check_status,
                record.ai_check_result or '',
                record.error_message or '',
                record.check_duration or 0
            ])
            
            # 设置数据行样式
            row_num = idx + 1
            for cell in ws[row_num]:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # 根据状态设置颜色
            status_cell = ws[f'C{row_num}']
            if record.check_status == 'success':
                status_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                status_cell.font = Font(color='006100')
            elif record.check_status == 'failed':
                status_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                status_cell.font = Font(color='9C0006')
        
        # 添加汇总信息
        ws.append([])
        ws.append(['汇总信息'])
        ws.append(['批次ID', batch_id])
        ws.append(['总数量', summary.total_count])
        ws.append(['成功数量', summary.success_count])
        ws.append(['失败数量', summary.failed_count])
        ws.append(['总耗时(秒)', summary.total_duration])
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def export_to_pdf(self, batch_id: str) -> bytes:
        """
        导出为PDF文件
        
        Args:
            batch_id: 批次ID
            
        Returns:
            PDF文件字节流
        """
        # 获取批次汇总
        summary = self.db.query(CheckSummary).filter(
            CheckSummary.batch_id == batch_id
        ).first()
        
        if not summary:
            raise ValueError(f"Batch {batch_id} not found")
        
        # 获取所有检查记录
        records = self.db.query(SQLCheckRecord).filter(
            SQLCheckRecord.batch_id == batch_id
        ).order_by(SQLCheckRecord.id).all()
        
        # 创建PDF文档
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=18,
        )
        
        # 构建内容
        elements = []
        
        # 样式
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1  # 居中
        )
        
        # 添加标题
        title = Paragraph(f"SQL检查报告 - {batch_id}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2 * inch))
        
        # 添加汇总信息表格
        summary_data = [
            ['批次ID', batch_id],
            ['总数量', str(summary.total_count)],
            ['成功数量', str(summary.success_count)],
            ['失败数量', str(summary.failed_count)],
            ['总耗时(秒)', str(summary.total_duration)],
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 4*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 0.3 * inch))
        
        # 添加检查记录表格
        table_data = [['编号', '检查状态', 'SQL语句', '检查结果/错误']]
        
        for idx, record in enumerate(records, start=1):
            result_text = record.ai_check_result if record.check_status == 'success' else record.error_message
            # 限制文本长度
            sql_text = record.original_sql[:100] + '...' if len(record.original_sql) > 100 else record.original_sql
            result_text = (result_text or '')[:100] + '...' if result_text and len(result_text) > 100 else (result_text or '')
            
            table_data.append([
                str(idx),
                record.check_status,
                sql_text,
                result_text
            ])
        
        records_table = Table(table_data, colWidths=[0.5*inch, 1*inch, 3.5*inch, 3.5*inch])
        records_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        
        elements.append(records_table)
        
        # 构建PDF
        doc.build(elements)
        output.seek(0)
        
        return output.getvalue()
